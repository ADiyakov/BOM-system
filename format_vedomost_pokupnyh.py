import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ===== НАСТРОЙКИ =====
INPUT_XLSX = "BOM_split.xlsx"   # твой файл после compress + split_name_to_supplydoc
INPUT_SHEET = "BOM_compressed"               # имя листа во входном файле

OUTPUT_XLSX = "Vedomost_pokupnyh.xlsx"

OUT_HEADERS = [
    "Наименование",
    "Код продукции",
    "Обозначение документа на поставку",
    "Поставщик",
    "Куда входит (обозначение)",
    "на изделие",
    "в комплекты",
    "на регулир.",
    "Всего",
    "Примечание",
]
# =====================

def clean(s) -> str:
    return re.sub(r"\s+", " ", ("" if s is None else str(s))).strip()

def is_blank_row(values) -> bool:
    return all(clean(v) == "" for v in values)

def is_total_qty_cell(v) -> bool:
    # Итоговая строка из compress_by_name: =REPT(...)&CHAR(10)&SUM(...)
    if v is None:
        return False
    if isinstance(v, str):
        t = v.strip()
        return t.startswith("=REPT(") or ("REPT(" in t) or ("____" in t)
    return False

def to_int_or_empty(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s == "":
        return ""
    try:
        return int(s)
    except Exception:
        return v  # на всякий

def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[INPUT_SHEET] if INPUT_SHEET in wb.sheetnames else wb.active

    in_headers = [clean(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(in_headers)}  # 0-based

    for need in ("Category", "Name_Clean", "SupplyDoc", "Module", "Qty", "Comment"):
        if need not in idx:
            raise RuntimeError(f"Во входном файле нет колонки '{need}'. Есть: {in_headers}")

    i_cat = idx["Category"]
    i_name = idx["Name_Clean"]
    i_sup  = idx["SupplyDoc"]
    i_mod  = idx["Module"]
    i_qty  = idx["Qty"]
    i_com  = idx["Comment"]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "ВП"
    out_ws.append(OUT_HEADERS)

    underline_font = Font(underline="single")
    wrap = Alignment(wrap_text=True)

    # индексы колонок в выходе (1-based)
    OUT_COL_NAME = 1
    OUT_COL_SUPPLYDOC = 3
    OUT_COL_MODULE = 5
    OUT_COL_QTY_ON_ITEM = 6
    OUT_COL_TOTAL = 9
    OUT_COL_COMMENT = 10

    last_category = None

    # Считываем все строки входа в список (чтобы удобно делать группировку)
    rows = [tuple(c.value for c in r) for r in ws.iter_rows(min_row=2)]
    n = len(rows)

    pos_idx = 0
    while pos_idx < n:
        r = rows[pos_idx]
        if r is None or is_blank_row(r):
            pos_idx += 1
            continue

        cat = clean(r[i_cat])
        name_clean = clean(r[i_name])
        qty_val = r[i_qty]

        # пропускаем входные "итоги" от compress_by_name (сами посчитаем)
        if is_total_qty_cell(qty_val):
            pos_idx += 1
            continue

        # если строка не старт позиции (Name_Clean пустой) — это мусор/обрыв,
        # просто пропустим (в норме такого не должно быть после compress)
        if name_clean == "":
            pos_idx += 1
            continue

        # 1) Заголовок категории при смене (перед позицией)
        if cat and cat != last_category:
            row_cat = [""] * len(OUT_HEADERS)
            row_cat[0] = cat
            out_ws.append(row_cat)
            out_ws.cell(out_ws.max_row, 1).font = underline_font

            out_ws.append([""] * len(OUT_HEADERS))  # пустая строка после категории
            last_category = cat

        # ---- собираем группу "позиция" ----
        group = []
        j = pos_idx
        while j < n:
            rr = rows[j]
            if rr is None or is_blank_row(rr):
                break

            # пропускаем входные итоги
            if is_total_qty_cell(rr[i_qty]):
                j += 1
                continue

            # новая позиция начинается с непустого Name_Clean
            if j != pos_idx and clean(rr[i_name]) != "":
                break

            group.append(rr)
            j += 1

        # ---- выводим строки группы ----
        first_out_row = None
        last_out_row = None

        for gi, gr in enumerate(group):
            g_name = clean(gr[i_name])
            g_supply = clean(gr[i_sup])
            g_module = clean(gr[i_mod])
            g_qty = to_int_or_empty(gr[i_qty])
            g_comment = clean(gr[i_com])

            out_row = [""] * len(OUT_HEADERS)

            if gi == 0:
                out_row[OUT_COL_NAME - 1] = g_name
                out_row[OUT_COL_SUPPLYDOC - 1] = g_supply
                out_row[OUT_COL_COMMENT - 1] = g_comment

            out_row[OUT_COL_MODULE - 1] = g_module
            out_row[OUT_COL_QTY_ON_ITEM - 1] = g_qty
            out_row[OUT_COL_TOTAL - 1] = g_qty  # в строках позиции "Всего" = Qty

            out_ws.append(out_row)

            if first_out_row is None:
                first_out_row = out_ws.max_row
            last_out_row = out_ws.max_row

        # ---- итог по позиции: только колонка "Всего" с чертой+SUM ----
        # Итоговую строку выводим только если в позиции больше одной строки (по твоей логике compress)
        if len(group) > 1 and first_out_row is not None and last_out_row is not None:
            total_row = [""] * len(OUT_HEADERS)
            col_letter = get_column_letter(OUT_COL_QTY_ON_ITEM)  # суммируем "на изделие"
            formula = f'=REPT("_",4)&CHAR(10)&SUM({col_letter}{first_out_row}:{col_letter}{last_out_row})'
            total_row[OUT_COL_TOTAL - 1] = formula
            out_ws.append(total_row)
            out_ws.cell(out_ws.max_row, OUT_COL_TOTAL).alignment = wrap

        # 3) Пустая строка ПОСЛЕ ПОЗИЦИИ (т.е. перед следующим Наименованием)
        out_ws.append([""] * len(OUT_HEADERS))

        pos_idx = j

    # ширины
    widths = [55, 14, 45, 22, 22, 10, 10, 10, 12, 35]
    for i, w in enumerate(widths, 1):
        out_ws.column_dimensions[get_column_letter(i)].width = w

    out_wb.save(OUTPUT_XLSX)
    print(f"OK: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
