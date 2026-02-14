# parse_spec_to_bom.py

from docx import Document
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ==========================
# НАСТРОЙКИ
# ==========================

INPUT_DOCX = "КОР_03_11_000_Кросс_модуль_4FTR_Спецификация.docx"          # входной Word
OUTPUT_XLSX = "BOM.xlsx"          # выходной Excel
MODULE_CODE = "КОР-03.11.000"     # код модуля


# ==========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==========================

def clean(text: str) -> str:
    """Убирает лишние пробелы"""
    return re.sub(r"\s+", " ", (text or "").strip())


quote_re = re.compile(r'["“”«»](.+?)["“”«»]')


def extract_manufacturer(text: str) -> str:
    """
    Производитель = последнее в кавычках
    """
    matches = quote_re.findall(text or "")
    return clean(matches[-1]) if matches else ""


# ==========================
# ОСНОВНОЙ ПАРСЕР
# ==========================

def parse_spec(input_path, module_code):
    doc = Document(input_path)

    rows = []

    current_section = None

    # Буферы одной позиции
    buf_pos = None
    buf_qty = None

    buf_name = []
    buf_comment = []
    buf_designation = []


    def flush():
        """
        Сохраняет текущую позицию в результат
        """
        nonlocal buf_pos, buf_qty
        nonlocal buf_name, buf_comment, buf_designation

        if buf_pos is None:
            return

        name = clean(" ".join(buf_name))
        comment = clean(" ".join(buf_comment))
        designation = clean(" ".join(buf_designation))

        manufacturer = extract_manufacturer(name)

        rows.append([
            module_code,
            current_section,
            int(buf_pos),
            name,
            manufacturer,
            designation,
            int(buf_qty),
            comment
        ])

        # Сброс буферов
        buf_pos = None
        buf_qty = None

        buf_name = []
        buf_comment = []
        buf_designation = []


    # ==========================
    # ПРОХОД ПО ТАБЛИЦАМ
    # ==========================

    for table in doc.tables:

        for row in table.rows:

            cells = [clean(c.text) for c in row.cells]

            if not any(cells):
                continue

            row_text = " ".join(cells)


            # --------------------------
            # Определение разделов
            # --------------------------

            if "Стандартные изделия" in row_text:
                flush()
                current_section = "Стандартные"
                continue

            if "Прочие изделия" in row_text:
                flush()
                current_section = "Прочие"
                continue

            if any(x in row_text for x in ["Документация", "Детали"]):
                flush()
                current_section = None
                continue


            # --------------------------
            # Пропуск заголовков
            # --------------------------

            if "Формат" in row_text and "Поз." in row_text:
                continue


            if current_section not in ("Стандартные", "Прочие"):
                continue


            # --------------------------
            # Нормализация колонок
            # --------------------------
            # Ожидаемый формат:
            # Формат | Зона | Поз | Обозн | Наим | Кол | Прим

            while len(cells) < 7:
                cells.append("")

            fmt, zone, pos_c, desig_c, name_c, qty_c, comm_c = cells[:7]


            # --------------------------
            # Новая позиция?
            # --------------------------

            is_new = (
                re.fullmatch(r"\d{1,3}", pos_c)
                and
                re.fullmatch(r"\d{1,4}", qty_c)
            )


            if is_new:
                flush()
                buf_pos = pos_c
                buf_qty = qty_c


            # --------------------------
            # Накопление колонок
            # --------------------------

            if buf_pos is not None:

                if desig_c:
                    buf_designation.append(desig_c)

                if name_c:
                    buf_name.append(name_c)

                if comm_c:
                    buf_comment.append(comm_c)


    # Последняя позиция
    flush()

    return rows


# ==========================
# ВЫГРУЗКА В EXCEL
# ==========================

def save_xlsx(rows, path):

    wb = Workbook()
    ws = wb.active

    ws.title = "BOM"

    ws.append([
        "Module",
        "Section",
        "Pos",
        "Name",
        "Manufacturer",
        "PartNumber",
        "Qty",
        "Comment"
    ])


    for r in rows:
        ws.append(r)


    widths = [16, 14, 6, 60, 28, 28, 6, 60]

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


    wb.save(path)


# ==========================
# ТОЧКА ВХОДА
# ==========================

if __name__ == "__main__":

    data = parse_spec(INPUT_DOCX, MODULE_CODE)

    data.sort(key=lambda x: (0 if x[1]=="Стандартные" else 1, x[2]))

    save_xlsx(data, OUTPUT_XLSX)

    print(f"Готово: {len(data)} позиций → {OUTPUT_XLSX}")
