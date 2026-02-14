# add_category.py
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_XLSX = "BOM_all_sorted.xlsx"                 # или BOM_all_sorted.xlsx
SHEET_NAME = "BOM"                          # если лист иначе — поменяй
OUTPUT_XLSX = "BOM_all_with_category.xlsx"  # выход

# --------------------------
# Категории (порядок важен)
# --------------------------
RULES = [
    # 1) Крепеж
    ("Крепежные изделия", re.compile(
        r"(?iu)\b(винт|болт|гайка|шайба|гровер|саморез|шуруп|закл[её]пк|шпильк|штифт|шплинт|стопорн)\b"
    #    r"|(?:\bDIN\b|\bISO\b|\bГОСТ\b)\s*\d+", # чтобы не перехватывать другие категории с указанными стандартами
    ,
        re.I
    )),

    # 2) Источники питания
    ("Источники питания (блоки питания из SV01)", re.compile(
        r"(?iu)\b(блок\s*питан|источник\s*питан|power\s*supply|psu|ac[-\s]*dc|dc[-\s]*dc)\b",
        re.I
    )),

    # 3) Предохранители (лучше раньше резисторов/прочего)
    ("Предохранители", re.compile(
        r"(?iu)\b(предохранител|fuse)\b|\bFU\d+\b",
        re.I
    )),

    # 4) Коммутирующие (разъемы/реле)
    ("Коммутирующие изделия (разъемы, реле)", re.compile(
        r"(?iu)\b(раз[ъе]м|вилка|розетка|клемм\w*|колодк\w*|панельк\w*|терминал\w*|разъём|din\s*41612|d[-\s]?sub|db[-\s]?\d+|header|socket)\b"
        r"|\b(реле|контактор|переключател\w*)\b",
        re.I
    )),
 
   # 10) Полупроводники (ставим ближе к концу, чтобы не перехватывать "панель DIP-32"). Нат, перенесли перед конструктивом, чтобы не реагировать на Корпус и т.п.
    ("Полупроводниковые изделия", re.compile(
    r"(?iu)\b("
    r"линейн\w*\s+регулятор|"
    r"регулятор\s+напряжени\w*\s+линейн\w*|"
    r"модул\w*\s+памят\w*|"
    r"мост\s+диодн\w*|"
    r"оптрон\w*|"
    r"позистор\w*|"
    r"стабилитрон\w*|"
    r"термистор\w*|"
    r"усилител\w*\s+звуков\w*\s+мощност\w*|"
    r"усилител\w*\s+низк\w*\s+частот\w*|"
    r"усилител\w*\s+операцион\w*|"
    r"операцион\w*\s+усилител\w*|"
    r"регистр|"
    r"резонатор|"
    r"инвертер|"
    r"микросхем\w*|"
    r"диод|"
    r"транзистор|"
    r"буфер|"
    r"контроллер|"
    r"оптопар\w*|"
    r"триггер|"
    r"счетчик|"
    r"микроконтроллер|"
    r"усилител\w*"
    r")\b"
    r"|\b(IC|DD|DA|VT|D)\d+\b"
    r"|PIC\d+|SN74|AM29|ispLSI|MC74|uPD|UPD",
    re.I
))
,

    # 5) Конструктив
    ("Конструктивные элементы (шкаф, стойки резьбовые, кронштейн шрофф, ручка, втулки, шасси из главной сп)", re.compile(
        r"(?iu)\b(шкаф|держатель|шасси|кронштейн|стойк\w*|стойка\s*резьбов\w*|втулк\w*|ручк\w*|корпус|кожух|панел\w*|крышк\w*|рам\w*|направляющ\w*|салазк\w*|планк\w*)\b"
        r"|Schroff|Assmann",
        re.I
    )),

    # 6) Моточные
    ("Моточные изделия (катушки, дроссели, трансформаторы)", re.compile(
        r"(?iu)\b(трансформатор|дроссел\w*|катушк\w*|индуктивн\w*|inductor|choke)\b|\bTR\d+\b",
        re.I
    )),

    # 7) Конденсаторы
    ("Конденсаторы", re.compile(
        r"(?iu)\b(конденсатор|capacitor)\b|\bC\d+\b",
        re.I
    )),

    # 8) Резисторы
    ("Резисторы", re.compile(
        r"(?iu)\b(резистор|сборка\s*резисторн\w*|resistor)\b|\bR\d+\b",
        re.I
    )),

    # 9) Органы управления/индикации
    ("Органы управления и индикации (светодиоды, лампы, кнопки)", re.compile(
        r"(?iu)\b(светодиод|ламп\w*|кнопк\w*|индикатор\w*|led)\b",
        re.I
    )),

 
]

DEFAULT_CATEGORY = "Прочие изделия"


def clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def classify(name: str) -> str:
    s = name or ""
    for cat, rx in RULES:
        if rx.search(s):
            return cat
    return DEFAULT_CATEGORY


def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    headers = [clean(str(c.value)) if c.value is not None else "" for c in ws[1]]
    if "Name" not in headers:
        raise RuntimeError(f"Нет колонки 'Name'. Заголовки: {headers}")

    name_col = headers.index("Name") + 1

    # Добавляем/находим колонку Category
    if "Category" in headers:
        cat_col = headers.index("Category") + 1
    else:
        cat_col = ws.max_column + 1
        ws.cell(1, cat_col).value = "Category"
        ws.column_dimensions[get_column_letter(cat_col)].width = 45

    # Заполняем
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        name = clean(str(name)) if name is not None else ""
        ws.cell(r, cat_col).value = classify(name)

    wb.save(OUTPUT_XLSX)
    print(f"OK: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
