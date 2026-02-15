# parse_specs_to_bom_many.py
# Основано на parse_spec_to_bom.py (актуальная версия пользователя)
#
# Модификации (по требованиям):
# - Pos может быть нечисловым: '-', '–', '—' -> считать валидной позицией и включать в вывод
# - При встрече секции "Материалы" парсинг файла прекращается (строки материалов не нужны для ВП)
# - В выходной файл добавлен столбец PosText (номер позиции или прочерк)
# - После каждого файла добавляется пустая строка (как было)

from docx import Document
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from specs_list import SPECS
from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent
SPECS_DIR = BASE_DIR / "specs"


# ==========================
# НАСТРОЙКИ
# ==========================

OUTPUT_XLSX = "BOMs_parsed.xlsx"  # один общий выходной Excel



# ==========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==========================

def clean(text: str) -> str:
    """Убирает лишние пробелы"""
    return re.sub(r"\s+", " ", (text or "").strip())


quote_re = re.compile(r'["“”«»](.+?)["“”«»]')


def extract_manufacturer(text: str) -> str:
    """Производитель = последнее в кавычках"""
    matches = quote_re.findall(text or "")
    return clean(matches[-1]) if matches else ""


# Прочерк в поле "Поз." бывает разный: -, – (en dash), — (em dash)
DASH_POS_RE = re.compile(r"^[\-–—]+$")


def is_pos_numeric(pos_c: str) -> bool:
    return bool(re.fullmatch(r"\d{1,3}", pos_c or ""))


def is_qty_numeric(qty_c: str) -> bool:
    return bool(re.fullmatch(r"\d{1,4}", qty_c or ""))


def is_pos_dash(pos_c: str) -> bool:
    return bool(DASH_POS_RE.fullmatch(pos_c or ""))


def pos_sort_key(pos_text: str) -> tuple[int, str]:
    """Ключ сортировки внутри блока: числа по возрастанию, прочерки в конце."""
    if is_pos_numeric(pos_text):
        return (0, f"{int(pos_text):03d}")
    # все нечисловые (в т.ч. прочерки) — после числовых, но стабильно
    return (1, pos_text or "")


# ==========================
# ОСНОВНОЙ ПАРСЕР
# ==========================

def parse_spec(input_path, module_code):
    doc = Document(input_path)

    rows = []
    current_section = None

    # Буферы одной позиции
    buf_pos_text = None     # строка: '71' или '-'/'—'/...
    buf_qty = None

    buf_name = []
    buf_comment = []
    buf_designation = []

    def flush():
        """Сохраняет текущую позицию в результат"""
        nonlocal buf_pos_text, buf_qty
        nonlocal buf_name, buf_comment, buf_designation

        if buf_pos_text is None:
            return

        name = clean(" ".join(buf_name))
        comment = clean(" ".join(buf_comment))
        designation = clean(" ".join(buf_designation))
        manufacturer = extract_manufacturer(name)

        # qty может быть пустым (если в исходнике съехало). Тут НЕ чиним.
        qty_val = int(buf_qty) if (buf_qty is not None and is_qty_numeric(buf_qty)) else 0

        rows.append([
            module_code,
            current_section,
            buf_pos_text,   # PosText: номер или прочерк
            name,
            manufacturer,
            designation,
            qty_val,
            comment
        ])

        # Сброс буферов
        buf_pos_text = None
        buf_qty = None
        buf_name = []
        buf_comment = []
        buf_designation = []

    stop_parsing = False

    for table in doc.tables:
        if stop_parsing:
            break

        for row in table.rows:
            cells = [clean(c.text) for c in row.cells]
            if not any(cells):
                continue

            row_text = " ".join(cells)

            # --------------------------
            # Остановка на секции "Материалы"
            # --------------------------
            # Как только встретили "Материалы" (обычно после Прочих изделий) — прекращаем парсинг файла.
            if "Материалы" in row_text:
                flush()
                stop_parsing = True
                break

            # --------------------------
            # Определение разделов
            # --------------------------
            if "Стандартные изделия" in row_text:
                flush()
                current_section = "Стандартные"
                continue

            if "Прочие изделия" in row_text:
                flush()
                current_section = "Прочие"
                continue

            if any(x in row_text for x in ["Документация", "Детали"]):
                flush()
                current_section = None
                continue

            # --------------------------
            # Пропуск заголовков
            # --------------------------
            if "Формат" in row_text and "Поз." in row_text:
                continue

            if current_section not in ("Стандартные", "Прочие"):
                continue

            # --------------------------
            # Нормализация колонок
            # --------------------------
            # Ожидаемый формат:
            # Формат | Зона | Поз | Обозн | Наим | Кол | Прим
            while len(cells) < 7:
                cells.append("")
            fmt, zone, pos_c, desig_c, name_c, qty_c, comm_c = cells[:7]

            # --------------------------
            # Новая позиция?
            # --------------------------
            # 1) Обычная позиция: Pos числовой И qty в этой же строке (как было)
            is_new_numeric = is_pos_numeric(pos_c) and is_qty_numeric(qty_c)

            # 2) Позиция-прочерк: Pos = '-'/'–'/'—' И qty в этой же строке
            #    (чтобы прочерки не склеивались друг с другом, считаем КАЖДУЮ такую строку началом новой позиции)
            is_new_dash = is_pos_dash(pos_c) and is_qty_numeric(qty_c) and bool(name_c)

            if is_new_numeric or is_new_dash:
                flush()
                buf_pos_text = pos_c
                buf_qty = qty_c

            # --------------------------
            # Накопление колонок
            # --------------------------
            if buf_pos_text is not None:
                if desig_c:
                    buf_designation.append(desig_c)
                if name_c:
                    buf_name.append(name_c)
                if comm_c:
                    buf_comment.append(comm_c)

    flush()
    return rows


# ==========================
# ВЫГРУЗКА В EXCEL (один файл)
# ==========================

def save_xlsx_many(specs, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # Заголовок
    ws.append([
        "Module",
        "Section",
        "PosText",   # <-- новый столбец
        "Name",
        "Manufacturer",
        "PartNumber",
        "Qty",
        "Comment"
    ])

    total_rows = 0

    for i, (input_docx, module_code) in enumerate(specs, start=1):
        #data = parse_spec(input_docx, module_code)
        input_path = SPECS_DIR / input_docx
        data = parse_spec(str(input_path), module_code)
        # сортировка блока: Section (Стандартные/Прочие) + PosText (числа, потом прочерки)
        data.sort(key=lambda x: (0 if x[1] == "Стандартные" else 1, pos_sort_key(x[2])))

        for r in data:
            ws.append(r)
            total_rows += 1

        # Пустая строка после каждого файла
        ws.append([""] * 8)

    # Ширины колонок
    widths = [16, 14, 8, 60, 28, 28, 6, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return total_rows


# ==========================
# ТОЧКА ВХОДА
# ==========================

if __name__ == "__main__":
    n = save_xlsx_many(SPECS, OUTPUT_XLSX)
    print(f"Готово: {len(SPECS)} файлов, {n} строк BOM → {OUTPUT_XLSX}")
