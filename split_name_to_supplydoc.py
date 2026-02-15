import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

INPUT_XLSX = "BOM_compressed_by_name.xlsx"
SHEET = "BOM"
OUTPUT_XLSX = "BOM_split.xlsx"

# --- patterns ---
VENDOR_QUOTED_AT_END = re.compile(r'\s*["“”«»]([^"“”«»]+)["“”«»]\s*$', re.U)

# DIN/ISO/ГОСТ/ТУ/ОСТ/СТО + номер(а)
STD_RE = re.compile(
    r'(?iu)\b('
    r'DIN|ISO|EN|IEC|ГОСТ(?:\s*Р)?|ТУ|ОСТ|СТО'
    r')\b'
    r'(?:\s*№?\s*[A-ZА-Я0-9][A-ZА-Я0-9\.\-\/]*|\s*\d[\d\.\-\/]*)*'
)

# Децимальные/КД обозначения: СЦМЕ.420009.001, АБВГ.666777.001 и т.п.
DECIMAL_RE = re.compile(r'(?iu)\b[А-ЯA-Z]{2,6}\.\d{3,6}\.\d{2,3}(?:\.\d{1,3})?\b')

# Плотный part number: длина >= 6, есть цифра, допускаем -,.,_
PART_TOKEN_RE = re.compile(r'(?iu)\b(?=[A-ZА-Я0-9\-\._]{6,}\b)(?=.*\d)[A-ZА-Я0-9][A-ZА-Я0-9\-\._]*\b')

# “служебные” токены (корпуса/диэлектрики и т.п.) — не считать part number
PART_BAD = set(map(str.upper, [
    "SOIC", "SOIC-8", "SOIC8", "DIP", "DIP-8", "DIP-32", "PLCC", "QFN", "SOT",
    "X7R", "NP0", "C0G"
]))

def clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def split_name(name: str):
    """
    Возвращает (Name_Clean, SupplyDoc)

    Требования:
    - Vendor попадает в конец SupplyDoc только если реально есть в кавычках в конце исходной строки.
    - Vendor в SupplyDoc в кавычках.
    - Перед vendor точка с запятой не ставится.
    """
    s = clean(name)

    # 1) Vendor в кавычках в конце — держим ОТДЕЛЬНО, НЕ кладём в supply_parts
    vendor = None
    m = VENDOR_QUOTED_AT_END.search(s)
    if m:
        vendor = clean(m.group(1))
        s = clean(s[:m.start()])  # удалили "Vendor" из анализируемой строки

    supply_parts = []  # сюда только стандарты/децимал/артикул

    # 2) Стандарты
    stds = [clean(x.group(0)) for x in STD_RE.finditer(s)]
    if stds:
        supply_parts.extend(stds)
        s = clean(STD_RE.sub("", s))

    # 3) Децимальные обозначения
    decs = [clean(x.group(0)) for x in DECIMAL_RE.finditer(s)]
    if decs:
        supply_parts.extend(decs)
        s = clean(DECIMAL_RE.sub("", s))

    # 4) Part token (берём “главный” ближе к концу)
    tokens = []
    for t in PART_TOKEN_RE.findall(s):
        if t.upper() in PART_BAD:
            continue
        # refdes вида R12, C5, D3 — не part number
        if re.fullmatch(r'(?iu)[RCVDT]{1,2}\d+', t):
            continue
        tokens.append(t)

    if tokens:
        part = tokens[-1]
        supply_parts.append(part)
        s = clean(re.sub(re.escape(part), "", s, count=1))

    # Уникализация (с сохранением порядка)
    supply_parts = list(dict.fromkeys([p for p in supply_parts if p]))

    # Финальная сборка SupplyDoc
    supply_doc = "; ".join(supply_parts)
    if vendor:
        supply_doc = (supply_doc + f' "{vendor}"') if supply_doc else f'"{vendor}"'

    name_clean = clean(s)
    return name_clean, supply_doc


def ensure_col(ws, headers, col_name, width=40):
    if col_name in headers:
        return headers.index(col_name) + 1
    col = ws.max_column + 1
    ws.cell(1, col).value = col_name
    ws.column_dimensions[get_column_letter(col)].width = width
    return col


def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active

    headers = [clean(str(c.value)) if c.value is not None else "" for c in ws[1]]
    if "Name" not in headers:
        raise RuntimeError(f"Нет колонки Name. Заголовки: {headers}")

    name_col = headers.index("Name") + 1
    name_clean_col = ensure_col(ws, headers, "Name_Clean", width=60)
    supply_col = ensure_col(ws, headers, "SupplyDoc", width=55)

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        name = "" if name is None else str(name)
        nc, sd = split_name(name)
        ws.cell(r, name_clean_col).value = nc or None
        ws.cell(r, supply_col).value = sd or None

    wb.save(OUTPUT_XLSX)
    print(f"OK: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
