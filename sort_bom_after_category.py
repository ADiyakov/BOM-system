import re
from pathlib import Path
from openpyxl import load_workbook, Workbook


# ========== НАСТРОЙКИ ==========
INPUT_XLSX = "BOM_with_category.xlsx"
SHEET_NAME = "BOM"
OUTPUT_XLSX = "BOM_with_category_sorted.xlsx"

# Названия категорий (как у вас в classify_category.py)
CAT_FASTENERS = "Крепежные изделия"
CAT_RESISTORS = "Резисторы"
CAT_CAPACITORS = "Конденсаторы"

SECTION_STANDARD = "Стандартные"
# ===============================


# --------- утилиты ----------
def s(x) -> str:
    return "" if x is None else str(x).strip()

def norm_space(x: str) -> str:
    return re.sub(r"\s+", " ", s(x)).strip()

def lower_ru_lat(x: str) -> str:
    # простая нормализация
    return norm_space(x).lower()

def natural_key(text: str):
    """Естественная сортировка: 'A2' < 'A10'."""
    t = lower_ru_lat(text)
    parts = re.split(r"(\d+)", t)
    out = []
    for p in parts:
        if p.isdigit():
            out.append(int(p))
        else:
            out.append(p)
    return tuple(out)

def to_float(num: str) -> float:
    return float(num.replace(",", "."))
# ----------------------------


# --------- крепёж ----------
DIN_RE = re.compile(r"(?iu)\bDIN\s*([0-9]{2,6})\b")
ISO_RE = re.compile(r"(?iu)\bISO\s*([0-9]{2,6})\b")
GOST_RE = re.compile(r"(?iu)\bГОСТ(?:\s*Р)?\s*([0-9]{2,6})\b")

M_SIZE_RE = re.compile(r"(?iu)\b[МM]\s*(\d+(?:[.,]\d+)?)\s*(?:[x×\*]\s*(\d+(?:[.,]\d+)?))?")
# пример: M3x10, М8×25, M12

def fastener_key(name: str):
    """
    Сортировка для крепежа:
    1) DIN/ISO/ГОСТ номер (если есть)
    2) алфавит по имени
    3) размеры M.. x L.. (резьба, длина)
    4) natural fallback
    """
    n = norm_space(name)

    din = DIN_RE.search(n)
    iso = ISO_RE.search(n)
    gost = GOST_RE.search(n)

    std_num = None
    std_rank = 9  # DIN раньше ISO раньше ГОСТ, но это необязательно; можно одинаково
    if din:
        std_num = int(din.group(1))
        std_rank = 1
    elif iso:
        std_num = int(iso.group(1))
        std_rank = 2
    elif gost:
        std_num = int(gost.group(1))
        std_rank = 3

    if std_num is None:
        std_rank = 9
        std_num = 10**9

    # размеры M
    m = M_SIZE_RE.search(n)
    thread = 10**9
    length = 10**9
    if m:
        try:
            thread = to_float(m.group(1))
        except Exception:
            thread = 10**9
        if m.group(2):
            try:
                length = to_float(m.group(2))
            except Exception:
                length = 10**9

    # алфавитная часть (без проброса цифр — оставим natural)
    alpha = lower_ru_lat(n)

    return (std_rank, std_num, alpha, thread, length, natural_key(n))
# ----------------------------


# --------- резисторы / конденсаторы ----------
# сопротивление: 10 Ом, 4,7кОм, 1МОм, 10k, 2.2M, 1R0 и т.п.
RES_RE = re.compile(
    r"(?iu)\b("
    r"\d+(?:[.,]\d+)?\s*(?:ом|к?ом|м?ом)"
    r"|\d+(?:[.,]\d+)?\s*(?:k|m)\b"
    r"|\d+r\d+"
    r")\b"
)
POWER_RE = re.compile(r"(?iu)\b(\d+(?:[.,]\d+)?)\s*(?:вт|w)\b")

CAP_RE = re.compile(
    r"(?iu)\b(\d+(?:[.,]\d+)?)\s*(пф|pf|нф|nf|мкф|µf|uf|uF|nF|pF)\b"
)
VOLT_RE = re.compile(r"(?iu)\b(\d+(?:[.,]\d+)?)\s*(?:в|v)\b")

def parse_res_ohms(name: str) -> float:
    n = lower_ru_lat(name).replace(" ", "")
    m = RES_RE.search(n)
    if not m:
        return 10**18
    token = m.group(1)

    # 1R0 формат
    if re.fullmatch(r"(?iu)\d+r\d+", token):
        token = token.lower().replace("r", ".")
        return to_float(token)

    # k/m суффиксы латиницей
    m2 = re.fullmatch(r"(?iu)(\d+(?:[.,]\d+)?)(k|m)", token)
    if m2:
        val = to_float(m2.group(1))
        mult = 1e3 if m2.group(2).lower() == "k" else 1e6
        return val * mult

    # ом/кОм/МОм по-русски (возможны "ком", "мом")
    m3 = re.fullmatch(r"(?iu)(\d+(?:[.,]\d+)?)(ом|ком|мом)", token)
    if m3:
        val = to_float(m3.group(1))
        unit = m3.group(2).lower()
        if unit == "ом":
            return val
        if unit == "ком":
            return val * 1e3
        if unit == "мом":
            return val * 1e6

    return 10**18

def parse_power_w(name: str) -> float:
    m = POWER_RE.search(lower_ru_lat(name))
    if not m:
        return 10**9
    try:
        return to_float(m.group(1))
    except Exception:
        return 10**9

def parse_cap_pf(name: str) -> float:
    """
    Возвращаем ёмкость в pF (для сортировки).
    """
    n = lower_ru_lat(name)
    m = CAP_RE.search(n)
    if not m:
        return 10**18
    val = to_float(m.group(1))
    unit = m.group(2).lower()

    if unit in ("пф", "pf"):
        return val
    if unit in ("нф", "nf"):
        return val * 1e3
    if unit in ("мкф", "µf", "uf"):
        return val * 1e6
    return 10**18

def parse_voltage_v(name: str) -> float:
    n = lower_ru_lat(name)
    m = VOLT_RE.search(n)
    if not m:
        return 10**9
    try:
        return to_float(m.group(1))
    except Exception:
        return 10**9
# --------------------------------------------


def row_sort_key(row_dict: dict):
    cat = s(row_dict.get("Category"))
    name = s(row_dict.get("Name"))
    section = s(row_dict.get("Section"))

    cat_key = lower_ru_lat(cat)

    # 1) крепёж или стандартные
    if cat == CAT_FASTENERS or section == SECTION_STANDARD:
        return (cat_key, 0, fastener_key(name))

    # 2) резисторы
    if cat == CAT_RESISTORS:
        r_ohm = parse_res_ohms(name)
        p_w = parse_power_w(name)
        return (cat_key, 1, r_ohm, p_w, natural_key(name))

    # 3) конденсаторы
    if cat == CAT_CAPACITORS:
        c_pf = parse_cap_pf(name)
        v_v = parse_voltage_v(name)
        return (cat_key, 2, c_pf, v_v, natural_key(name))

    # 4) прочее
    return (cat_key, 9, natural_key(name))


def main():
    in_path = Path(INPUT_XLSX)
    if not in_path.exists():
        raise FileNotFoundError(f"Не найден {INPUT_XLSX}")

    wb = load_workbook(in_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    headers = [s(c.value) for c in ws[1]]
    if "Category" not in headers or "Name" not in headers:
        raise RuntimeError(f"Ожидались колонки Category и Name. Есть: {headers}")

    # читаем строки
    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        row_dict = {headers[i]: values[i] for i in range(len(headers))}
        rows.append((row_dict, values))

    # сортируем
    rows_sorted = sorted(rows, key=lambda rv: row_sort_key(rv[0]))

    # пишем в новый файл, сохраняя заголовок
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title

    out_ws.append(headers)
    for _, values in rows_sorted:
        out_ws.append(values)

    out_wb.save(OUTPUT_XLSX)
    print(f"OK: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
