from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ====== НАСТРОЙКИ ======
INPUT_XLSX  = "Vedomost_pokupnyh.xlsx"
SHEET_NAME  = "ВП"
OUTPUT_XLSX = "Vedomost_pokupnyh_wrapped.xlsx"

# Колонки ведомости (1-based):
# 1 Наименование
# 2 Код продукции
# 3 Обозначение документа на поставку
# 4 Поставщик
# 5 Куда входит (обозначение)
# 6 на изделие
# 7 в комплекты
# 8 на регулир.
# 9 Всего
# 10 Примечание
MAX_COL = 10

NAME_COL = 1
SUPPLYDOC_COL = 3
MODULE_COL = 5
QTY_COL = 6
TOTAL_COL = 9
COMMENT_COL = 10

LINE_LEN = 4  # длина "____" в итоговой ячейке

# Ширины колонок (как в образце ТГТ-01.00.00 ВП, сдвинуто под нашу структуру)
COLUMN_WIDTHS = {
    1: 38.77734375,  # Наименование
    2: 20.33203125,  # Код продукции
    3: 33.0,         # Обозначение документа на поставку
    4: 27.44140625,  # Поставщик
    5: 36.33203125,  # Куда входит (обозначение)
    6: 8.5546875,    # на изделие
    7: 8.5546875,    # в комплекты
    8: 8.5546875,    # на регулир.
    9: 12.77734375,  # Всего
    10: 12.77734375, # Примечание
}

# Лимиты длины строки "в символах" для переноса по словам
MAX_CHARS = {
    NAME_COL: 38,       # Наименование
    SUPPLYDOC_COL: 33,  # Обозначение документа на поставку
    COMMENT_COL: 12,    # Примечание (узко в образце)
}
# =======================


def is_blank_cell(v) -> bool:
    return v is None or str(v).strip() == ""


def is_blank_row(ws, r) -> bool:
    for c in range(1, MAX_COL + 1):
        if not is_blank_cell(ws.cell(r, c).value):
            return False
    return True


def set_col_widths(ws):
    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def words_wrap(text: str, max_len: int) -> list[str]:
    """
    Перенос по словам.
    Длинные "слова" (артикулы) НЕ режем — остаются как есть.
    """
    s = "" if text is None else str(text).strip()
    if s == "":
        return []
    parts = s.split()
    lines = []
    cur = ""
    for w in parts:
        if cur == "":
            cur = w
        elif len(cur) + 1 + len(w) <= max_len:
            cur += " " + w
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


def collect_col_text(ws, start, end, col) -> str:
    """Собираем текст по всей позиции из колонки col (склейка через пробел)."""
    parts = []
    for rr in range(start, end):
        v = ws.cell(rr, col).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return " ".join(parts).strip()


def is_total_cell(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return ("_" * LINE_LEN) in s or "____" in s or "REPT(" in s or s.startswith("=")


def find_total_row(ws, start, end):
    for rr in range(start, end):
        if is_total_cell(ws.cell(rr, TOTAL_COL).value):
            return rr
    return None


def sum_qty_in_position(ws, start, end) -> int:
    total = 0
    for rr in range(start, end):
        v = ws.cell(rr, QTY_COL).value
        if v is None:
            continue
        try:
            total += int(str(v).strip())
        except Exception:
            pass
    return total


def collect_comments_by_module(ws, start, end, total_row):
    """
    Собираем примечания ПО КАЖДОМУ MODULE-блоку внутри позиции.
    Module-блок: от строки с непустым Module до строки перед следующим Module или total_row/конца позиции.
    Примечания внутри блока склеиваем через пробел.
    Возвращает список блоков:
      [{"start": row_idx, "end": row_idx_exclusive, "text": "..."}]
    """
    blocks = []
    cur = None  # текущий блок

    for rr in range(start, end):
        if total_row is not None and rr == total_row:
            # Итог — не часть module-блока
            break

        mod = ws.cell(rr, MODULE_COL).value
        mod_str = "" if mod is None else str(mod).strip()

        if mod_str != "":
            # старт нового блока
            if cur is not None:
                blocks.append(cur)
            cur = {"start": rr, "end": rr + 1, "parts": []}

        if cur is not None:
            # собираем comment с текущей строки (включая “продолжения” пока не встретили новый Module)
            v = ws.cell(rr, COMMENT_COL).value
            if v is not None:
                s = str(v).strip()
                if s:
                    cur["parts"].append(s)
            cur["end"] = rr + 1

    if cur is not None:
        blocks.append(cur)

    # финализация текста
    out = []
    for b in blocks:
        text = " ".join(b["parts"]).strip()
        out.append({"start": b["start"], "end": b["end"], "text": text})
    return out


def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    set_col_widths(ws)

    r = 2  # пропускаем заголовок

    while r <= ws.max_row:
        if is_blank_row(ws, r):
            r += 1
            continue

        # позиция = блок строк до пустой строки
        start = r
        end = r
        while end <= ws.max_row and not is_blank_row(ws, end):
            end += 1
        # позиция: [start, end)

        total_row = find_total_row(ws, start, end)

        # === 1) Перенос общих полей позиции: Наименование и SupplyDoc ===
        name_text = collect_col_text(ws, start, end, NAME_COL)
        supply_text = collect_col_text(ws, start, end, SUPPLYDOC_COL)

        name_lines = words_wrap(name_text, MAX_CHARS[NAME_COL])
        supply_lines = words_wrap(supply_text, MAX_CHARS[SUPPLYDOC_COL])

        needed_lines_common = max(1, len(name_lines) if name_lines else 1, len(supply_lines) if supply_lines else 1)
        existing_lines = end - start

        if needed_lines_common > existing_lines:
            insert_at = end
            ws.insert_rows(insert_at, amount=(needed_lines_common - existing_lines))
            end += (needed_lines_common - existing_lines)

            # после вставки total_row смещается, если он был внутри позиции ниже insert_at
            if total_row is not None and total_row >= insert_at:
                total_row += (needed_lines_common - existing_lines)

        # очищаем и записываем 1 и 3 по строкам позиции
        for rr in range(start, end):
            ws.cell(rr, NAME_COL).value = None
            ws.cell(rr, SUPPLYDOC_COL).value = None

        for i, line in enumerate(name_lines):
            ws.cell(start + i, NAME_COL).value = line
        for i, line in enumerate(supply_lines):
            ws.cell(start + i, SUPPLYDOC_COL).value = line

        # === 2) Примечание: перенос по каждому Module-блоку ===
        # сначала соберём блоки (по текущим границам start/end/total_row)
        comment_blocks = collect_comments_by_module(ws, start, end, total_row)

        # очистим колонку Примечание во всей позиции (кроме пустой строки-разделителя)
        for rr in range(start, end):
            ws.cell(rr, COMMENT_COL).value = None

        # Чтобы вставки строк не сдвигали ещё не обработанные блоки,
        # обрабатываем блоки СНИЗУ ВВЕРХ.
        for b in reversed(comment_blocks):
            text = b["text"]
            if not text:
                continue

            lines = words_wrap(text, MAX_CHARS[COMMENT_COL])
            need = max(1, len(lines))
            have = b["end"] - b["start"]

            if need > have:
                insert_at = b["end"]  # вставляем в конец блока (до следующего Module/итога)
                ws.insert_rows(insert_at, amount=(need - have))

                # сдвигаем end позиции
                end += (need - have)

                # если итог ниже — он тоже сдвигается
                if total_row is not None and total_row >= insert_at:
                    total_row += (need - have)

                # также нужно скорректировать start/end всех блоков ВЫШЕ (но мы идём снизу вверх — не нужно)

            # пишем строки примечания начиная с b["start"]
            for i, line in enumerate(lines):
                ws.cell(b["start"] + i, COMMENT_COL).value = line

        # === 3) Итог "Всего": пересчёт числом (без формул) ===
        if total_row is not None:
            s = sum_qty_in_position(ws, start, end)
            ws.cell(total_row, TOTAL_COL).value = ("_" * LINE_LEN) + "\n" + str(s)

        # переход к следующей позиции
        r = end + 1

    wb.save(OUTPUT_XLSX)
    print(f"OK: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
