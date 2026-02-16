import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ========= НАСТРОЙКИ =========
INPUT_XLSX = "BOM_with_category_sorted.xlsx"
INPUT_SHEET = "BOM"

OUTPUT_XLSX = "BOM_compressed_by_name.xlsx"
OUTPUT_SHEET = "BOM_compressed"

NAME_COL_HEADER = "Name"
KEEP_HEADERS = {"Module", "PosText", "Qty"}  # эти поля не чистим в строках-повторах
QTY_HEADER = "Qty"

# "черта" над числом в ячейке итога (REPT("_", N))
LINE_LEN = 4
# ============================


def normalize_header(v):
    return (str(v).strip() if v is not None else "")


def main():
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb[INPUT_SHEET] if INPUT_SHEET in wb.sheetnames else wb.active

    headers = [normalize_header(c.value) for c in ws[1]]
    if NAME_COL_HEADER not in headers:
        raise RuntimeError(f"Нет колонки {NAME_COL_HEADER}. Заголовки: {headers}")
    if QTY_HEADER not in headers:
        raise RuntimeError(f"Нет колонки {QTY_HEADER}. Заголовки: {headers}")

    col = {h: i + 1 for i, h in enumerate(headers)}  # 1-based
    name_col = col[NAME_COL_HEADER]
    qty_col = col[QTY_HEADER]

    # Считываем данные (без заголовка). Пустые строки сохраняем как есть.
    data = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        data.append(row_vals)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = OUTPUT_SHEET

    # Заголовок
    out_ws.append(headers)

    wrap = Alignment(wrap_text=True)

    out_row_idx = 1  # текущая последняя заполненная строка в out_ws (заголовок = 1)
    i = 0
    while i < len(data):
        row = data[i]
        name = row[name_col - 1]

        # Если строка без Name — копируем как есть и идём дальше
        if name is None or str(name).strip() == "":
            out_ws.append(row)
            out_row_idx += 1
            i += 1
            continue

        # Граница блока по одинаковому Name (строгое совпадение)
        j = i
        while j < len(data):
            n2 = data[j][name_col - 1]
            if n2 is None or str(n2).strip() == "":
                break
            if str(n2) != str(name):
                break
            j += 1

        group_size = j - i

        # Запоминаем диапазон строк по Qty в выходном листе для суммы
        first_group_out = out_row_idx + 1  # первая строка группы будет следующей вставкой

        # Пишем строки группы
        for k in range(i, j):
            rvals = list(data[k])

            # Для повторов (кроме первой строки блока) чистим всё, кроме KEEP_HEADERS
            if k != i:
                for h, ci in col.items():
                    if h in KEEP_HEADERS:
                        continue
                    rvals[ci - 1] = None

            out_ws.append(rvals)
            out_row_idx += 1

        last_group_out = out_row_idx

        # Итоговую строку добавляем только если строк в блоке больше 1
        if group_size > 1:
            sum_row = [None] * len(headers)
            qty_letter = get_column_letter(qty_col)
            # черта + перенос строки + SUM по диапазону Qty внутри группы
            formula = f'=REPT("_",{LINE_LEN})&CHAR(10)&SUM({qty_letter}{first_group_out}:{qty_letter}{last_group_out})'
            sum_row[qty_col - 1] = formula
            out_ws.append(sum_row)
            out_row_idx += 1
            out_ws.cell(out_row_idx, qty_col).alignment = wrap

        # Пустая строка после каждого блока — ВСЕГДА
        out_ws.append([None] * len(headers))
        out_row_idx += 1

        i = j

    # Чуть удобочитаемые ширины (если такие колонки есть)
    widths = {
        "Module": 16,
        "Section": 14,
        "PosText": 8,
        "Category": 22,
        "Name": 60,
        "Manufacturer": 28,
        "PartNumber": 28,
        "Qty": 8,
        "Comment": 60,
        "SupplyDoc": 55,
        "Name_Clean": 60,
    }
    for h, w in widths.items():
        if h in col:
            out_ws.column_dimensions[get_column_letter(col[h])].width = w

    out_wb.save(OUTPUT_XLSX)
    print(f"OK: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
